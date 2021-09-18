import openpyxl as pxl
from dateutil import relativedelta
import xlwings as xw
from source import Bean
import os
from datetime import datetime
start = datetime.now()

class Account:
    def __init__(self, file_path):
        self.file_path = file_path
        self.fetch_data()
        self.calc_intr()
    
    def fetch_data(self):
        wb = pxl.load_workbook(self.file_path)
        sheet = wb.active
        asset_list = list()
        for row in sheet.iter_rows(min_row=2,min_col=1,values_only=True):
            ast = Bean.Asset(*row)
            asset_list.append(ast)
        self.pf = Bean.Portfolio(asset_list)
    
    def calc_intr(self):
        count = 0
        start = 0
        file_name = "XIRRcalc.xlsx"
        workbook = pxl.Workbook()
        sheet = workbook.active
        for i in self.pf.asset_list:
            if i.sip == "N" or i.sip== "n":
                t = i.time_inv[2] + i.time_inv[1]*0.0833 + i.time_inv[0]*0.00274
                intr = (i.a_val/i.inv_amt)**(1/t) - 1
                i.int_rate = intr
            else:
                
                for j in range(i.time_inv[2]*12 + i.time_inv[1]+1):
                    sheet["A{}".format(start+j+1)] = i.a_date + relativedelta.relativedelta(months=j)
                    sheet["B{}".format(start+j+1)] = -i.a_inv
                sheet["A{}".format(start+j+2)] = datetime.now()
                sheet["B{}".format(start+j+2)] = i.a_val
                count += 1
                sheet["C{}".format(count)] = "=XIRR(B{}:B{},A{}:A{})".format(start+1,start+j+2,start+1,start+j+2)
                start += j+2
                
        workbook.save(filename=file_name)    
        ex_app = xw.App(visible=False)
        workbook = ex_app.books.open(file_name)
        k=1
        for i in self.pf.asset_list:
            if i.sip in ["Y","y"]:
                i.int_rate = workbook.sheets["Sheet"].range('C{}'.format(k)).value
                k += 1
        workbook.close()
        ex_app.quit()
        os.remove(file_name)



    